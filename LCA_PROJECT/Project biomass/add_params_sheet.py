from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Config workbook lives next to this script (LCA_PROJECT/Project biomass/).
PATH = str(Path(__file__).resolve().parent / "LCA_process_config.xlsx")
wb = openpyxl.load_workbook(PATH)

# ── Style helpers ────────────────────────────────────────────────────────────
thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(cell, fill=None, bold=False, size=10, color="000000", align="left", wrap=False):
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BORDER

HDR  = PatternFill("solid", fgColor="1F4E79")
SEC1 = PatternFill("solid", fgColor="2E75B6")
SEC2 = PatternFill("solid", fgColor="375623")
SEC3 = PatternFill("solid", fgColor="7B3F00")
SEC4 = PatternFill("solid", fgColor="833C00")
SEC5 = PatternFill("solid", fgColor="4A235A")   # purple — syngas balance
ROW1 = PatternFill("solid", fgColor="DDEEFF")
ROW2 = PatternFill("solid", fgColor="E2EFDA")
ROW3 = PatternFill("solid", fgColor="FCE4D6")
ROW4 = PatternFill("solid", fgColor="FFF2CC")
ROW5 = PatternFill("solid", fgColor="F3E8FF")   # light purple — syngas balance

# ═══════════════════════════════════════════════════════════════════════════
# SHEET: Parameters
# ═══════════════════════════════════════════════════════════════════════════
if "Parameters" in wb.sheetnames:
    del wb["Parameters"]

ws = wb.create_sheet("Parameters", 1)
ws.freeze_panes = "A3"

COLS = ["section", "parameter_name", "symbol", "default_value",
        "formula", "unit", "source", "used_in_phase", "description"]

for c, h in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    sc(cell, fill=HDR, bold=True, color="FFFFFF", size=11)
ws.row_dimensions[1].height = 22

rows = [
    # A – User Config
    ("A – User Config", "BIOMASS_NAME",       "—",      "Poplar 4",  "—",
     "—",                    "USER CONFIG cell",                  "All",
     "Matched against 'subcategory' column in Biomass_Selection sheet"),
    ("A – User Config", "d1",                 "d1",     50,          "—",
     "km",                   "Biomass_Selection col H",           "Ph3 – Transport",
     "Transport distance: feedstock to pyrolysis plant"),
    ("A – User Config", "d2",                 "d2",     100,         "—",
     "km",                   "Biomass_Selection col I",           "Ph3 – Transport",
     "Transport distance: bio-oil to port / end use"),
    ("A – User Config", "d3",                 "d3",     150,         "—",
     "km",                   "Biomass_Selection col J",           "Ph3 – Transport",
     "Transport distance: bio-char to soil amendment site"),
    ("A – User Config", "k_el",               "k_el",   0.15,        "—",
     "kWh / kg dry biomass", "Biomass_Selection col K",           "Ph2 – Pyrolysis conversion",
     "Electricity intensity for pyrolysis. Default: Bridgwater (2012) fast pyrolysis."),
    ("A – User Config", "k_th",               "k_th",   0.75,        "—",
     "MJ / kg dry biomass",  "Biomass_Selection col L",           "Ph2 – Pyrolysis conversion",
     "Process heat intensity for pyrolysis. Default: Bridgwater (2012) fast pyrolysis."),
    ("A – User Config", "MC_ITERATIONS",       "—",      100,         "—",
     "iterations",           "USER CONFIG cell",                  "Monte Carlo",
     "Number of Monte Carlo iterations. Use 500-1000 for publication quality."),
    ("A – User Config", "KEEP_FOREGROUND",     "—",      "True",      "—",
     "bool",                 "USER CONFIG cell",                  "Cleanup",
     "Default is now True. Each run creates a biomass-specific DB (greenfuel_foreground_{slug}) "
     "so multiple runs coexist without overwriting each other."),
    ("A – User Config", "FG_DB_NAME",          "—",      "(auto)",    "greenfuel_foreground_{biomass_slug}",
     "str",                  "Computed in Step 5",                 "All foreground steps",
     "Brightway DB name for this run's foreground activity. Auto-generated from BIOMASS_NAME "
     "so each biomass gets its own DB. Avoids overwriting previous runs."),

    # B – Physical Constants
    ("B – Physical Constants", "C_BIO_FRAC",      "C_bio",   0.60, "—",
     "kg C / kg bio-oil",   "IPCC / Bridgwater (2012)",          "Ph4 – Use (marine)",
     "Carbon mass fraction in produced bio-oil (daf). Used to compute biogenic CO2 emission."),
    ("B – Physical Constants", "HHV_SYNGAS_MJ_KG","HHV_syn", 12.0, "—",
     "MJ / kg syngas",      "Bridgwater (2012)",                  "Ph5 – Co-product credits",
     "Higher heating value of pyrolysis syngas. Range 10-14 MJ/kg in literature."),
    ("B – Physical Constants", "ETA_COMB",         "eta_comb",0.85, "—",
     "—",                   "Bridgwater (2012)",                  "Ph5 – Co-product credits",
     "Syngas combustion efficiency for heat substitution credit."),
    ("B – Physical Constants", "WATER_KG_PER_KG",  "w_water", 0.05, "—",
     "kg / kg dry biomass", "Assumption",                         "Ph2 – Pyrolysis conversion",
     "Process water per kg dry biomass. Covers quenching and scrubbing."),

    # C – Yield Parameters
    ("C – Yield Parameters", "oil_yield_kg_per_kg",    "x_raw", "(yield file)", "—",
     "kg / kg as-received",
     "Biomass_Selection col D / yield_results.xlsx",   "Step 2 – Parameter extraction",
     "Bio-oil yield as-received basis (fraction). Primary LCA sensitivity parameter."),
    ("C – Yield Parameters", "bio_oil_hhv_mj_per_kg",  "T",     "(yield file)", "—",
     "MJ / kg bio-oil",
     "Biomass_Selection col E / yield_results.xlsx",   "Step 2 – Parameter extraction",
     "HHV of produced bio-oil. Used in FU normalisation: z = 1/T, y = z/x_raw."),
    ("C – Yield Parameters", "temperature_c",            "T_py",  "(yield file)", "—",
     "degC",
     "Biomass_Selection col F / yield_results.xlsx",   "Step 2 – Parameter extraction",
     "Optimal pyrolysis temperature for selected biomass. Linked to yield and k_el / k_th."),
    ("C – Yield Parameters", "feedstock_moisture_pct_ar","MC",    "(yield file)", "—",
     "wt%",
     "Biomass_Selection col G / yield_results.xlsx",   "Step 2 – Parameter extraction",
     "Feedstock moisture content as-received. Informative — not directly used in exchange amounts."),
    ("C – Yield Parameters", "char_yield_kg_per_kg",    "c_raw", "(yield file)", "—",
     "kg / kg as-received",
     "yield_results.xlsx (not in Biomass_Selection)",  "Step 2 – Parameter extraction",
     "Char yield as-received. Used to derive c = c_raw x y (char per FU)."),
    ("C – Yield Parameters", "gas_yield_kg_per_kg",     "g_raw", "(yield file)", "—",
     "kg / kg as-received",
     "yield_results.xlsx (not in Biomass_Selection)",  "Step 2 – Parameter extraction",
     "Gas yield as-received. Used to derive g = g_raw x y (syngas per FU)."),

    # D – Derived Parameters
    ("D – Derived (Script)", "y",          "y",        "(computed)", "1 / (x_raw x T)   OR   kg_biomass_per_mj_bio_oil_hhv",
     "kg biomass / FU",  "Computed in Step 2",  "Ph1  Ph2  Ph3",
     "Biomass required to produce 1 MJ usable bio-oil (the functional unit)."),
    ("D – Derived (Script)", "z",          "z",        "(computed)", "1 / T",
     "kg bio-oil / FU",  "Computed in Step 2",  "Ph3  Ph4  Ph5",
     "Bio-oil mass per FU. Reference product of foreground activity."),
    ("D – Derived (Script)", "c",          "c",        "(computed)", "c_raw x y",
     "kg bio-char / FU", "Computed in Step 2",  "Ph3  Ph5",
     "Bio-char mass per FU. Used for transport (Ph3) and soil credit (Ph5)."),
    ("D – Derived (Script)", "g",          "g",        "(computed)", "g_raw x y",
     "kg syngas / FU",   "Computed in Step 2",  "Ph5 – Syngas energy balance",
     "Syngas mass per FU. Used in syngas energy balance to compute credited heat."),
    ("D – Derived (Script)", "tkm1",       "tkm1",     "(computed)", "d1 x (y / 1000)",
     "tonne-km / FU",    "Computed in Step 2",  "Ph3 – Feedstock transport",
     "Transport work: feedstock to pyrolysis plant."),
    ("D – Derived (Script)", "tkm2",       "tkm2",     "(computed)", "d2 x (z / 1000)",
     "tonne-km / FU",    "Computed in Step 2",  "Ph3 – Bio-oil transport",
     "Transport work: bio-oil to port / end use."),
    ("D – Derived (Script)", "tkm3",       "tkm3",     "(computed)", "d3 x (c / 1000)",
     "tonne-km / FU",    "Computed in Step 2",  "Ph3 – Bio-char transport",
     "Transport work: bio-char to soil amendment site."),
    ("D – Derived (Script)", "el_proc",    "el_proc",  "(computed)", "k_el x y",
     "kWh / FU",         "Computed in Step 5",  "Ph2 – Electricity exchange",
     "Electricity consumed per FU. Amount used in p_elec exchange. No syngas electricity credit in current model scope."),
    ("D – Derived (Script)", "heat_proc",  "heat_proc","(computed)", "k_th x y",
     "MJ / FU",          "Computed in Step 5",  "Ph2 – Process heat exchange",
     "Internal heat demand of pyrolysis per FU. Also used as the CAP for syngas credit (see below)."),
    ("D – Derived (Script)", "water_proc", "w_proc",   "(computed)", "w_water x y",
     "kg / FU",          "Computed in Step 5",  "Ph2 – Process water exchange",
     "Process water consumed per FU. Amount used in p_water exchange."),
    ("D – Derived (Script)", "co2_kg",     "co2_kg",   "(computed)", "z x C_bio x (44 / 12)",
     "kg CO2 / FU",      "Computed in Step 5",  "Ph4 – Biogenic CO2 emission",
     "Biogenic CO2 from marine combustion of bio-oil. Biosphere exchange."),

    # ── SECTION E: Syngas Energy Balance (new) ───────────────────────────────
    ("E – Syngas Balance", "gross_syngas_heat_mj", "q_syn_gross", "(computed)",
     "g x HHV_syn x eta_comb",
     "MJ / FU",  "Computed in Step 5",  "Ph5 – Syngas energy balance",
     "Total heat that syngas COULD deliver if fully combusted. Upper bound only — not directly used as credit."),
    ("E – Syngas Balance", "credited_syngas_heat_mj", "q_syn_credit", "(computed)",
     "min(q_syn_gross, heat_proc)  =  min(g x HHV_syn x eta_comb,  k_th x y)",
     "MJ / FU",  "Computed in Step 5",  "Ph5 – Syngas credit exchange (negative technosphere)",
     "Heat actually credited as avoided burden. CAPPED at internal heat demand (heat_proc). "
     "Syngas offsets only what the reactor would otherwise buy from the natural gas market. "
     "If syngas output exceeds heat demand, the surplus is NOT credited."),
    ("E – Syngas Balance", "surplus_syngas_heat_mj", "q_syn_surplus", "(computed)",
     "max(q_syn_gross - heat_proc, 0)  =  max(g x HHV_syn x eta_comb - k_th x y, 0)",
     "MJ / FU",  "Computed in Step 5",  "Ph5 – Not modelled",
     "Surplus syngas energy beyond internal heat demand. Intentionally NOT credited in this model. "
     "Reason: surplus syngas is not assumed to displace any external energy within the system boundary. "
     "Could be revisited if syngas export to grid is modelled explicitly."),
    ("E – Syngas Balance", "NOTE — electricity credit", "—", "DISABLED", "—",
     "—",  "Computed in Step 5",  "Ph5 – Not modelled",
     "Syngas electricity substitution is intentionally disabled in the current model scope. "
     "Placeholder exists in code for future extension."),
]

section_styles = {
    "A – User Config":        (SEC1, ROW1),
    "B – Physical Constants": (SEC2, ROW2),
    "C – Yield Parameters":   (SEC3, ROW3),
    "D – Derived (Script)":   (SEC4, ROW4),
    "E – Syngas Balance":     (SEC5, ROW5),
}

prev_sec = None
row_idx = 2
for r in rows:
    sec = r[0]
    sf, rf = section_styles[sec]
    if sec != prev_sec:
        prev_sec = sec
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(row=row_idx, column=c,
                           value=sec if c == 1 else "")
            sc(cell, fill=sf, bold=True, color="FFFFFF", size=10)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
    for c, v in enumerate(r, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        sc(cell, fill=rf,
           bold=(c == 1),
           wrap=(c in (5, 9)),
           align="center" if c in (3, 4, 6) else "left")
    ws.row_dimensions[row_idx].height = 32 if r[4] != "—" else 18
    row_idx += 1

col_widths = [26, 30, 12, 16, 38, 22, 38, 28, 62]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
# UPDATE LCI_Mapping — insert symbol + formula-with-symbols columns
# ═══════════════════════════════════════════════════════════════════════════
ws_lci = wb["LCI_Mapping"]
ws_lci.insert_cols(7, 2)

HDR2 = PatternFill("solid", fgColor="1F4E79")
for col, title in [(7, "parameter_symbol"), (8, "amount_formula (symbols)")]:
    cell = ws_lci.cell(row=1, column=col, value=title)
    sc(cell, fill=HDR2, bold=True, color="FFFFFF", size=11)

formula_map = {
    "Feedstock (hardwood)":               ("y",        "y"),
    "Feedstock alt. (softwood)":          ("y",        "y"),
    "Electricity":                        ("el_proc",  "k_el x y"),
    "Process heat":                       ("heat_proc","k_th x y"),
    "Process water":                      ("w_proc",   "w_water x y"),
    "Feedstock transport":                ("tkm1",     "d1 x (y / 1000)"),
    "Bio-oil transport":                  ("tkm2",     "d2 x (z / 1000)"),
    "Bio-char transport":                 ("tkm3",     "d3 x (c / 1000)"),
    "Marine combustion proxy":            ("z",        "z  =  1 / T"),
    "Syngas heat credit (avoided)":       ("q_syn_credit", "-min(g x HHV_syn x eta_comb,  k_th x y)"),
    "Bio-char soil credit (avoided)":     ("c",        "-c  =  -(c_raw x y)"),
    "Feedstock (nut shells / hulls)":     ("y",        "y"),
    "Feedstock alt. (fruit/olive residues)": ("y",     "y"),
    "Feedstock (mill/wood residues)":     ("y",        "y"),
    "Feedstock alt. (fruit/seed residues)":  ("y",     "y"),
}

phase_fills_lci = {
    "Ph1": PatternFill("solid", fgColor="F2F2F2"),
    "Ph2": PatternFill("solid", fgColor="FFFFFF"),
    "Ph3": PatternFill("solid", fgColor="F2F2F2"),
    "Ph4": PatternFill("solid", fgColor="FFFFFF"),
    "Ph5": PatternFill("solid", fgColor="FFF2CC"),
}

for row in ws_lci.iter_rows(min_row=2, max_row=ws_lci.max_row):
    label     = row[2].value
    phase_key = str(row[1].value or "")[:3]
    rf        = phase_fills_lci.get(phase_key, PatternFill("solid", fgColor="FFFFFF"))
    sym, formula = formula_map.get(str(label) if label else "", ("—", "—"))

    c7 = ws_lci.cell(row=row[0].row, column=7, value=sym)
    c8 = ws_lci.cell(row=row[0].row, column=8, value=formula)
    sc(c7, fill=rf, bold=True, align="center")
    sc(c8, fill=rf)

ws_lci.column_dimensions["G"].width = 20
ws_lci.column_dimensions["H"].width = 30

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(PATH)
print("Done.")
print("Sheets:", wb.sheetnames)
